import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import os
from kneed import KneeLocator
from sklearn.linear_model import LinearRegression
from openpyxl import workbook #pip install openpyxl
from openpyxl import load_workbook
from collections import deque
directory='C:/Users/SamSung/Desktop/OneDrive_2020-06-17/6_PDMS_Donuts/cleanerdata'
#filepath='C:/Users/SamSung/Desktop/OneDrive_2020-06-17/6_PDMS_Donuts/cleanerdata/1.xlsx'
directory2="C:/Users/SamSung/Desktop/OneDrive_2020-06-17/6_PDMS_Donuts/"

def get_stiffness():
    nr=6
    nc=1
    datalist=[]
    idx=1
    idy=1
    stiffness=[]
    ylow=-10
    yhigh=-2
    x=deque(os.listdir(directory))
    x.popleft()
    x.rotate(-3)
    x.appendleft('1.xlsx')
    print(x)
    for filename in x:

        filepath=os.path.join(directory,filename)
        data=pd.read_excel(filepath).drop(['Unnamed: 0'],axis=1)
        X=data.iloc[:,0].values.reshape(-1,1).ravel()
        Y=data.iloc[:,1].values.reshape(-1,1).ravel()
        kneedle=KneeLocator(X,Y,S=1,curve='concave',direction='increasing',interp_method="interp1d")
        xthreshold=round(kneedle.knee,3)
        data=data[data.Disp<xthreshold]
        data=data[data.Load<yhigh]
        data=data[data.Load>ylow]
        #-> cuts out below threshold
        data=data[data.index < data['Load'].idxmin()]#separate two curves
        X=data.iloc[:,0].values.reshape(-1,1)
        Y=data.iloc[:,1].values.reshape(-1,1).ravel()
        lr=LinearRegression()
        model=lr.fit(X,Y)
        prediction=model.predict(X)
        print("X len is",X.shape)
        stiffness.append(lr.coef_[0])
        #plt.plot(X,prediction,color='red')
        #plt.plot(X,Y)
        #plt.plot(X,lr.coef_[0][0]*X+lr.intercept_[0],'green')

        #-> gets slopes
        datalist.append(data)


    fig,axes=plt.subplots(nr,nc,sharex=True)
    for ax in axes:
        #print(ax)
        ax.plot(datalist[idx-1].Disp,datalist[idx-1].Load,color='red')#dry
        ax.plot(datalist[idx].Disp,datalist[idx].Load)#hydrated
        idx+=2
    plt.show()

    print(stiffness)
    # write stiffness values in xlsx file
    # have to change cell location everytime!
    column=4
    for x in range(len(stiffness)):
        wb = load_workbook(directory2+"stiffness1.xlsx")
        ws = wb['Sheet1']
        ws.cell(2+x,column).value=stiffness[x]
        ws.cell(14,column).value=ylow
        ws.cell(15,column).value=yhigh
        wb.save(directory2+"stiffness1.xlsx")
    return

get_stiffness()
